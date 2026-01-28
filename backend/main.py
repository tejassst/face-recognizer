from urllib.request import Request
from fastapi import HTTPException
from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from deepface import DeepFace
from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path
import logging
import os
import warnings

# Suppress TensorFlow warnings and force CPU
os.environ['CUDA_VISIBLE_DEVICES'] = '-1'
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
warnings.filterwarnings('ignore')

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


KNOWN_FACES_DIR = "../assets/known_faces"
UPLOADS_DIR = "uploads"
EXCEL_FILE = "../assets/log.xlsx"

# Create directories if they don't exist
os.makedirs(KNOWN_FACES_DIR, exist_ok=True)
os.makedirs(UPLOADS_DIR, exist_ok=True)


app = FastAPI(
    title="Face Recognition API",
    description="API for face recognition mobile app",
    version="1.0.0"
)

# CORS - Allow React Native app to connect
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    logger.info("Face Recognition API starting up...")
    logger.info("Running on CPU mode (GPU disabled)")

@app.get("/")
def read_root():
    return {
        "message": "Face Recognition API",
        "status": "running",
        "version": "1.0.0"
    }

@app.get("/health")
def health_check():
    return {"status": "healthy"}

@app.post("/recognize")
async def recognize_face(file: UploadFile = File(...)):
    """
    This function:
    1. Receives an uploaded photo
    2. Saves it temporarily
    3. Searches database for matching face
    4. Return the person's face or unknown
    """
    temp_path = None

    try: 
        logger.info(f"Received image: {file.filename}")
        temp_path = os.path.join(UPLOADS_DIR, f"temp_{file.filename}")

        with open(temp_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content) 

        logger.info(f"Written to file! and saved to :{temp_path}")
        known_faces = [
            f for f in os.listdir(KNOWN_FACES_DIR)
            if f.endswith(('.jpg','.png','.jpeg'))
        ]
        if not known_faces:
            return {
                "error": "No faces in database",
                "status": "empty_database",
                "message": "Please add faces to the database first"
            }
        logger.info(f"Searching database ({len(known_faces)} faces)...")
        
        # mtcnn - best for accessories, ssd - fast, opencv - very lenient
        detectors = ['mtcnn', 'ssd', 'opencv']
        result = None
        used_detector = None
        
        for detector in detectors:
            try:
                logger.info(f"🔍 Trying detector: {detector}")
                result = DeepFace.find(
                    img_path=temp_path,
                    db_path=KNOWN_FACES_DIR,
                    model_name="Facenet512",  # Best accuracy: 98.4%
                    detector_backend=detector,
                    enforce_detection=True,
                    silent=True
                )
                used_detector = detector
                logger.info(f"✅ Success with {detector}")
                break  # Found a face, stop trying other detectors
            except Exception as e:
                logger.warning(f"❌ {detector} failed: {str(e)[:100]}")
                continue

        if result is None:
            return {
                "error": "No face detected",
                "status": "no_face_detected",
                "message": "Could not detect a face in the image. Please try with better lighting."
            }

        if len(result) > 0 and not result[0].empty:
            matched_face = result[0].iloc[0]['identity']
            distance = result[0].iloc[0]['distance']
            # Distance: 0 = identical, 1+ = different
            # Convert to confidence: 0-1 (1 = 100% confident)
            confidence = max(0, 1 - (distance / 2))

            name = Path(matched_face).stem #.stem gets the filename without extension
            
            logger.info(f"✅ Match Found: {name} (confidence: {confidence:.2%}, detector: {used_detector})")
            return {
                "status": "match_found",
                "name": name,
                "confidence": round(confidence, 2),
                "distance": round(distance, 4),
                "detector_used": used_detector,
                "message": f"Recognized as {name}"
            }
        else:
            return {
                "status": "no_match_found",
                "message": "No match found in database",
                "detector_used": used_detector
            }
    except Exception as e:
        logger.error(f"❌ Error recognizing face: {e}")
        return {
            "error": "Error recognizing face",
            "status": "error",
            "message": str(e)
        }
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)
            
@app.get("/database")
def list_database():
    """
    This function: 
    1. Lists all image files in known_faces folder
    2. Extracts names from filenames
    3. Returns count and list of names
    """
    known_faces = [
            f for f in os.listdir(KNOWN_FACES_DIR)
            if f.endswith(('.jpg','.png','.jpeg'))
    ]
    if not known_faces:
        return {
            "total": 0,
            "names": [],
            "message": "Database is empty"
        }
    logger.info(f"Database has {len(known_faces)} face(s)")
    return {
        "total":len(known_faces),
        "names": sorted([Path(f).stem for f in known_faces])
    }

@app.post("/add-face")
async def add_face(file: UploadFile = File(...), name: str = "unknown"):
    """
    Adds a new face to the database
    Requires image file and person's name
    """
    temp_path = None
    try: 
        logger.info(f"📝 Adding face for: {name}")
        if not name or name == "unknown":
            raise HTTPException(status_code=400, detail="Please provide a valid name")
        
        temp_path = os.path.join(UPLOADS_DIR, f"temp_{file.filename}")
        with open(temp_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Use only built-in detectors (no extra packages needed)
        detectors = ['mtcnn', 'ssd', 'opencv']
        faces_detected = False
        used_detector = None
        
        for detector in detectors:
            try:
                logger.info(f"🔍 Trying detector: {detector}")
                faces = DeepFace.extract_faces(
                    img_path=temp_path,
                    detector_backend=detector,
                    enforce_detection=True
                )
                used_detector = detector
                logger.info(f"✅ {detector} detected {len(faces)} face(s)")
                faces_detected = True
                break
            except Exception as e:
                logger.warning(f"❌ {detector} failed: {str(e)[:100]}")
                continue
        
        if not faces_detected:
            return {
                "error": "No face detected",
                "status": "no_face",
                "message": "Please upload an image with a clear face."
            }
        
        file_path = os.path.join(KNOWN_FACES_DIR, f"{name}.jpg")
        os.rename(temp_path, file_path)
        temp_path = None

        logger.info(f"✅ Face added successfully for {name} using {used_detector}")
        return {
            "status": "success",
            "message": f"Face added successfully for {name}",
            "name": name,
            "detector_used": used_detector
        }
    except Exception as e:
        logger.error(f"❌ Error adding face: {e}")
        return {
            "error": "Error adding face",
            "status": "error",
            "message": str(e)
        }
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)
            
@app.post("/log-scan")
async def log_scan(request: Request):
    data = await request.json()
    name = data.get("name","Unknown")
    confidence = data.get("confidence", 0)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Confidence", "Timestamp"])
        wb.save(EXCEL_FILE)
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, confidence, timestamp])
    wb.save(EXCEL_FILE)
    return {"status": "success", "message": "Log saved successfully"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")